import csv

from functools import cached_property

import openpyxl

from io import BytesIO

from anyio import Path, to_thread
from loguru import logger

from app.exceptions import SheetException, SystemException
from app.exceptions import exception_handler
from app.utils.backupper import backup_manager
from app.utils.config import CONFIG


class BasicSheets:
    books: str = "books"
    users: str = "users"
    borrowed_books: str = "borrowed books"


class BasicCsvFiles:
    @cached_property
    def base_path(self) -> Path:
        return CONFIG.get_path("Directories", "csvfiles")

    @property
    def books(self) -> Path:
        return self.base_path / "books.csv"

    @property
    def users(self) -> Path:
        return self.base_path / "users.csv"

    @property
    def borrowed_books(self) -> Path:
        return self.base_path / "borrowed_books.csv"


csv_files = BasicCsvFiles()


class SheetConfig:
    _wbook = None
    _path_to_book = None

    @classmethod
    async def init(cls):
        cls._path_to_book = CONFIG.get_path("Files", "basicxlsxfile")
        try:
            cls._wbook = await to_thread.run_sync(openpyxl.load_workbook, cls._path_to_book)
            logger.info(f"[SheetConfig] Excel-файл загружен: {cls._path_to_book}")
        except FileNotFoundError:
            logger.warning(f"[SheetConfig] Excel-файл не найден. Создание нового: {cls._path_to_book}")
            cls._wbook = await to_thread.run_sync(cls._init_book)
        except Exception as e:
            logger.error(f"[SheetConfig] Ошибка загрузки Excel-файла: {e}")
            raise SheetException("Ошибка при открытии Excel-файла")

    @classmethod
    def _init_book(cls) -> openpyxl.Workbook:
        logger.info("[SheetConfig] Инициализация нового Excel-файла")
        wb = openpyxl.Workbook()
        del wb["Sheet"]

        books = wb.create_sheet(BasicSheets.books)
        books.append(["Название", "Авторство", "Публикация", "Категория", "ISBN", "Количество", "ID"])

        users = wb.create_sheet(BasicSheets.users)
        users.append(["ФИО", "Класс", "ID", "Дата создания"])

        borrowed_books = wb.create_sheet(BasicSheets.borrowed_books)
        borrowed_books.append(
            ["ID", "ID пользователя", "ID книги", "Название книги", "Имя пользователя", "Количество", "Дата выдачи"])

        wb.save(cls._path_to_book)
        logger.info(f"[SheetConfig] Новый Excel-файл создан: {cls._path_to_book}")
        return wb

    @classmethod
    @exception_handler(exception_cls=SheetException, message="Не удалось создать файлы csv")
    async def init_csv_files(cls):
        base_path = Path(CONFIG.get_path("Directories", "csvfiles"))
        await base_path.mkdir(parents=True, exist_ok=True)

        csv_definitions = {
            "books.csv": ["Название", "Авторство", "Публикация", "Категория", "ISBN", "Количество", "ID"],
            "users.csv": ["ФИО", "Класс", "ID", "Дата создания"],
            "borrowed_books.csv": ["ID", "Название книги", "Имя пользователя", "Количество", "Дата выдачи"]
        }

        for filename, headers in csv_definitions.items():
            path = base_path / filename
            if not await path.exists() or (await path.stat()).st_size == 0:
                async with await path.open("w", encoding="utf-8", newline="") as f:
                    writer = csv.writer(f)
                    await writer.writerow(headers)
                logger.info(f"[SheetConfig] CSV-файл создан/инициализирован: {path}")

    @classmethod
    def get_workbook(cls):
        if cls._wbook is None:
            raise SheetException("Не удалось найти рабочую таблицу excel")
        return cls._wbook

    @classmethod
    def get_path(cls):
        return cls._path_to_book


class XlsSheetController:
    @exception_handler(exception_cls=SystemException, message="Не удалось инициализировать контроллер xls")
    def __init__(self):
        # Листы теперь не кэшируются в self.*
        pass

    @property
    def wbook(self):
        """Всегда берём актуальный workbook из SheetConfig."""
        return SheetConfig.get_workbook()

    @property
    def books(self):
        return self.wbook[BasicSheets.books]

    @property
    def users(self):
        return self.wbook[BasicSheets.users]

    @property
    def borrowed_books(self):
        return self.wbook[BasicSheets.borrowed_books]

    async def update_wbook(self, content: bytes):
        stream = BytesIO(content)
        new_wbook = openpyxl.load_workbook(stream)

        # Подменяем глобальный workbook в SheetConfig
        SheetConfig._wbook = new_wbook

        # Сохраняем новый объект на диск синхронно в отдельном потоке
        await to_thread.run_sync(new_wbook.save, SheetConfig.get_path())
        await to_thread.run_sync(new_wbook.close)

    @exception_handler(exception_cls=SheetException, message="Ошибка при добавлении записи в таблицу Xlsx")
    async def append_row_to_sheet(self, sheet: str, row):
        def _append():
            self.wbook[sheet].append(row)
        await to_thread.run_sync(_append)

    async def append_many_rows_to_sheet(self, sheet: str, row_list: list[list]):
        def _append_many():
            for row in row_list:
                self.wbook[sheet].append(row)
            logger.debug(f"строки добавлены в {sheet}")
        await to_thread.run_sync(_append_many)

    @exception_handler(exception_cls=SheetException, message="Ошибка при очистке данных Excel-файла")
    def clear_all_sheets_except_headers(self):
        for sheet_name in [BasicSheets.books, BasicSheets.users, BasicSheets.borrowed_books]:
            sheet = self.wbook[sheet_name]
            max_row = sheet.max_row
            if max_row > 1:
                sheet.delete_rows(2, max_row - 1)
            logger.info(f"[XlsSheetController] Очищен лист '{sheet_name}', кроме заголовка")

    @exception_handler(exception_cls=SystemException, message="Ошибка при экспорте Excel в CSV")
    async def excel_to_csv(self):
        mapping = {
            BasicSheets.books: csv_files.books,
            BasicSheets.users: csv_files.users,
            BasicSheets.borrowed_books: csv_files.borrowed_books,
        }

        for sheet_name, csv_path in mapping.items():
            logger.info(f"[XlsSheetController] Начало экспорта листа '{sheet_name}' → {csv_path}")
            sheet = self.wbook[sheet_name]
            data = [list(row) for row in sheet.iter_rows(values_only=True)]

            # Резервное копирование CSV перед перезаписью
            await backup_manager.backup_csv_files_in_dir()

            async with await csv_path.open("w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                await writer.writerow(data)

            logger.info(f"[XlsSheetController] Лист '{sheet_name}' экспортирован в CSV: {csv_path}")



class SheetContextManager:
    def __init__(self):
        self.controller = None

    async def __aenter__(self):
        await SheetConfig.init()
        self.controller = XlsSheetController()
        return self.controller

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        try:
            await backup_manager.backup_xlsx_file()
            SheetConfig.get_workbook().save(SheetConfig.get_path())
            SheetConfig.get_workbook().close()
            logger.info("[SheetContextManager] Excel-файл сохранён и закрыт.")
        except:
            raise SheetException("Ошибка при завершении работы с Excel-файлом")
